import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .extraction import extract_table_titles, capture_change_idx
from database_interact.queries import extract_data_db
import json
import os

def concatenate_dataframes(df_list, limit = 3):
    def _add_empty_columns(df_list, limit=3):
        """
        Agrupa DataFrames en listas, cada una sera una fila, y añade columnas vacías según un límite.
        Esta funcion aun mantiene los dataframes en una lista, sin concatenarlos.

        Args:
            df_list (list): Lista de DataFrames.
            limit (int): Número máximo de DataFrames por grupo.

        Returns:
            list: Lista de filas de DataFrames agrupados.
        """
        current_group = []
        grouped_dataframes = []
        limit_counter = 0

        for i, df_i in enumerate(df_list):
            df:pd.DataFrame = df_i.copy()      # Copia el DataFrame original para evitar modificarlo directamente
            df.reset_index(drop=True, inplace=True)     # Reinicia el índice del DataFrame
            limit_counter += 1

            if i < len(df_list)-1:

                if limit_counter < limit:
                    df['empty'] = None
                
                else:
                    limit_counter = 0
                    current_group.append(df)
                    grouped_dataframes.append(current_group)
                    current_group = []
                    continue
                
                current_group.append(df)
            
            else:
                current_group.append(df)
                grouped_dataframes.append(current_group)
        
        return grouped_dataframes

    def _concatenate_column(row_packet_list):
        row_list = []

        for row_pack in row_packet_list:
            result = pd.concat(row_pack, axis=1, ignore_index=True)  # Apila las columnas formando una fila
            row_list.append(result)

        return row_list

    def _concatenate_row(rows_list):
        empty_row = pd.DataFrame({col:[None] for col in rows_list[0].columns})  # Crea una fila vacía con las mismas columnas que el primer DataFrame
        with_empty_row = []
        for i, df in enumerate(rows_list):
            with_empty_row.append(df)
            if i < len(rows_list) - 1:
                with_empty_row.append(empty_row)
                
        result = pd.concat(with_empty_row, ignore_index=True)  # Apila las filas
        return result

    rows_pack = _add_empty_columns(df_list, limit)
    rows_list = _concatenate_column(rows_pack)
    df_master = _concatenate_row(rows_list)
    return df_master

def generate_news_df(cuadros):
    new_df_list=[]

    def generate_df(date:str):
        date_sign=date+'::D'if 'DOMINGO' in date else date+'::O'
        columns_text=[['', date_sign, 'O', 'N']]
        new_df = pd.DataFrame(columns_text)
        return new_df
    
    def add_data(df, data):
        index = len(df)
        new_row_content = [index]
        for dat in data:
            if dat is not None:
                new_row_content.append(dat)
            else:
                new_row_content.append('-')
        df.loc[index] = new_row_content

    for cuadro in cuadros:
        titles_cuadro = extract_table_titles(cuadro, 1, complete=True)
        new_df = generate_df(titles_cuadro[0])

        for title in titles_cuadro[1:]:
            data_curr = extract_data_db(title)
            if data_curr:
                add_data(new_df, data_curr)
            else:
                raise ValueError(f'No se encontró el himno: {title}, en la base de datos.')
        new_df_list.append(new_df)

    return new_df_list

def load_config() ->dict:
    file = os.path.join('file_procces', 'config.json')
    with open(file, mode='r') as config:
        conf = json.load(config)
    return conf

def formating(df_master:pd.DataFrame, title_page:str):
    file = os.path.join('file_procces', f'{title_page}.xlsx')
    idx = capture_change_idx(df_master)

    df_master.to_excel(file, index=False, header=False)
    wb = load_workbook(file)
    ws = wb.active

    conf = load_config()

    style_new = Font(bold=True, color=conf['cl_new'])
    style_transpose = Font(bold=True, color=conf['cl_transpose'])
    fill_red = PatternFill(start_color=conf['fill_red'], end_color=conf['fill_red'], fill_type="solid")
    fill_green = PatternFill(start_color=conf['fill_green'], end_color=conf['fill_green'], fill_type="solid")
    fill_sunday = PatternFill(start_color=conf['fill_sunday'], end_color=conf['fill_sunday'], fill_type="solid")
    fill_other_day = PatternFill(start_color=conf['fill_other_day'], end_color=conf['fill_other_day'], fill_type="solid")
    alg_center = Alignment(horizontal = 'center')

    general_size = Font(size=conf['general_size'])
    header_size = conf['header_size']

    styles = {'new':style_new, 'transpose':style_transpose}
    fills = {'red':fill_red, 'green':fill_green, 'sunday':fill_sunday, 'other_day':fill_other_day}

    for row_idx in range(1, ws.max_row+1):
        height = 0
        derivation = row_idx % 8

        if derivation in (0, 1):
            height = conf['row_headers']

            if derivation == 1:
                for cell in ws[row_idx]:
                    cell.font = Font(size=header_size, bold=True)
                    cell.alignment = alg_center
    
        elif derivation in (2, 3, 4, 5, 6, 7):
            height = conf['row_general']

            for cell in ws[row_idx]:
                cell.font = general_size

        ws.row_dimensions[row_idx].height = height
        
    for col_idx in range(1, ws.max_column+1):
        width = 0
        letter = get_column_letter(col_idx)
        derivation = col_idx % 5

        if derivation == 1:
            width = conf['col_idx']

            for column in ws.iter_cols(min_col=col_idx, max_col=col_idx):
                for cell in column:
                    cell.alignment = alg_center
        
        elif derivation == 2:
            width = conf['col_title']
        
        elif derivation in (3, 4):
            width = conf['col_numbers']

            for column in ws.iter_cols(min_col=col_idx, max_col=col_idx):
                for i, cell in enumerate(column,1):
                    cell.font = Font(size=header_size, bold=cell.font.bold)
                    cell.alignment = alg_center
        
        elif derivation == 0:
            width = conf['col_space']
    
        ws.column_dimensions[letter].width = width

    for key, style in styles.items():
        for i, j in idx[key]:
            ws.cell(row=i+1, column=j+1).font += style

    for key, fill in fills.items():
        for i, j in idx[key]:
            ws.cell(row=i+1, column=j+1).fill = fill

    ws.insert_rows(1, amount=2)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ws.max_column)
    ws.cell(row=1, column=1, value=title_page).font=Font(name=conf['style_name'], bold=True, size=conf['size_name'])
    ws.cell(row=1, column=1).alignment=Alignment(horizontal='center')
    ws.row_dimensions[2].height = conf['distance_name']

    wb.save(file)
    print(f"Formato aplicado exitosamente en {file}.") 

def main():
    pass

if __name__ =="__main__":
    main()
